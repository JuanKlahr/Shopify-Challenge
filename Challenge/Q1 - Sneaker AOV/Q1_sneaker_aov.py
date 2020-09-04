# Juan Sebastian Klahr
# Q1 of Winter 2021 Data Science Intern Challenge
# Sept 2 2020

from openpyxl import load_workbook
from statistics import mean, median
from numpy import quantile


class Customer():
    def __init__(self):
        self.user_id = None
        self.shop_id = None
        self.amount_spent = 0
        self.total_items = 0
        self.total_orders = 0


def create_value_col(col):
    """Creates a tuple of values based off a tuple of cells"""
    values = []

    for row in col:  # Checks every cell in row
        for cell in row:  # Checks every cell in column
            values.append(cell.value)  # Adds cell value to new tuple

    return values


def mean_col(col):
    """Gets the mean of a column"""
    value_col = create_value_col(col)  # Converts the column to a tuple of values
    return round(mean(value_col), 2)


def median_col(col):
    """Gets the median of a column"""
    value_col = create_value_col(col)  # Converts the column to a tuple of values
    return round(median(value_col), 2)


def identify_big_customer(ws, row_num, order_amount, upper_limit, customer_lst):
    """Identifies who makes large orders"""
    # ID Columns
    shop_col = "B"
    user_col = "C"
    items_col = "E"

    #  Outputs info about large spender
    if order_amount > upper_limit:
        # Temporary IDs and item totals
        user_id = ws[user_col + str(row_num)].value
        shop_id = ws[shop_col + str(row_num)].value
        items_in_order = ws[items_col + str(row_num)].value
        new_customer = True

        #  Checks if costumer is new or a repeat big customer
        for customer in customer_lst:
            if shop_id == customer.shop_id and user_id == customer.user_id:
                new_customer = False
                #  Customer is repeat big customer, add new order to totals
                customer.amount_spent += order_amount
                customer.total_items += items_in_order
                customer.total_orders += 1

        #  Customer is new, start tracking them
        if new_customer:
            big_customer = Customer()
            big_customer.shop_id = shop_id
            big_customer.user_id = user_id
            big_customer.amount_spent = order_amount
            big_customer.total_items = items_in_order
            big_customer.total_orders += 1
            customer_lst.append(big_customer)


def create_big_customers_ws(wb, customer_lst):
    """Creates a new sheet that lists info about big customers"""
    bc_ws = wb.create_sheet("Big Customers")

    #  Column headers
    bc_ws["A1"] = "Shop ID"
    bc_ws["B1"] = "User ID"
    bc_ws["C1"] = "Total Orders"
    bc_ws["D1"] = "Total Spent"
    bc_ws["E1"] = "Total Items"

    #  Inputting call big customer info onto new Big Customers sheet
    for index, row in enumerate(bc_ws.iter_rows(min_row=2, max_row=len(customer_lst))):
        customer = customer_lst[index]
        row_num = str(index + 2)  # Add 2 because index starts at 0, but first row is row 2

        bc_ws["A" + row_num].value = customer.shop_id
        bc_ws["B" + row_num].value = customer.user_id
        bc_ws["C" + row_num].value = customer.total_orders
        bc_ws["D" + row_num].value = customer.amount_spent
        bc_ws["E" + row_num].value = customer.total_items


def check_outliers(col, wb):
    """Checks if there are any outliers in the column to determine if mean or median should be used"""
    value_col = create_value_col(col)  # Converts the column to a tuple of values
    data_ws = wb["Sheet1"]

    # Creating quartiles and fences for theoretical box plot
    lower_quartile = quantile(value_col, .25)
    upper_quartile = quantile(value_col, .75)
    iqr = upper_quartile - lower_quartile
    lower_fence = lower_quartile - iqr * 1.5
    upper_fence = upper_quartile + iqr * 1.5
    big_customers = []
    outliers = False

    # If a variable does not fall between the fences, variable is outlier
    for index, num in enumerate(value_col):
        if num < lower_fence or num > upper_fence:
            identify_big_customer(data_ws, index + 2, num, upper_fence, big_customers)  # Identify who the outliers are
            outliers = True

    if outliers:  # There were outliers
        print("Due to there being outliers in the data, Median is the best metric to use to find the AOV\n")
        print("New sheet listing Big Customers added to spreadsheet")
        create_big_customers_ws(wb, big_customers)
        return True
    else:  # There were no outliers
        print("There are no outliers that skew the distribution. Mean is best metric to use to find the AOV")
        return False


if __name__ == "__main__":
    filename = "2019 Winter Data Science Intern Challenge Data Set.xlsx"

    # Error handling on if source file is not in folder with program
    try:
        wb = load_workbook(filename)
        ws1 = wb["Sheet1"]

        #  Creates new Big Customers sheet if one already exists
        if "Big Customers" in wb.sheetnames:
            del wb["Big Customers"]

        col_letter = "D"  # Order Value column

        order_amount_col = ws1[col_letter + '2':col_letter + str(ws1.max_row)]  # Retrieving needed column

        print("Mean of the values in column {0} is {1}".format(col_letter, mean_col(order_amount_col)))
        print("Median of the values in column {0} is {1}".format(col_letter, median_col(order_amount_col)))

        check_outliers(order_amount_col, wb)
        wb.save(filename)

        input("\nPress Enter key to close program...")
    except FileNotFoundError:
        print("Source data file not found. Please read readme.txt")
        input("\n Press Enter key to close program...")
        wb = None
        ws1 = None
        exit()
    except:
        print("Unknown error. Please read readme.txt")
        input("\n Press Enter key to close program...")
        wb = None
        ws1 = None
        exit()
